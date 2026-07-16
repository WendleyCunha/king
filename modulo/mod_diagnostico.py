# =============================================================
# modulo/mod_diagnostico.py
# Diagnóstico N2 — Mapeamento de Atividades Operacionais (Backoffice)
#
# Segue a metodologia do documento "Mapeamento de Atividades
# Operacionais — Célula N2 (Backoffice)": inventário → organograma
# real (+ RACI preliminar) → Diário de Bordo → entrevistas → Gemba →
# Matriz de Atividades x Responsabilidades x Tempo → Mapa de Jornada
# de Trabalho por atividade → validação.
#
# Persistência: Firestore, reaproveitando get_db() de database.py.
# Cada seção (inventário, organograma, diário, etc.) é UM documento
# na coleção "diagnostico", com o conteúdo inteiro no campo "dados" —
# suficiente para o volume de uma equipe de backoffice e mais simples
# de manter do que um documento por linha.
#
# Entry point: renderizar_diagnostico(papel, user) — mesmo padrão de
# renderizar_rastreio / renderizar_tickets / renderizar_cartas.
# =============================================================
import json
import re
from datetime import date, datetime, time as dtime, timedelta, timezone
from io import BytesIO

import pandas as pd
import streamlit as st

from database import get_db, pode_editar

BRT = timezone(timedelta(hours=-3))


# ─────────────────────────────────────────────────────────────
# Persistência (Firestore) — coleção "diagnostico", 1 doc por seção
# ─────────────────────────────────────────────────────────────
def _ler(chave: str, default=None):
    doc = get_db().collection("diagnostico").document(chave).get()
    if doc.exists:
        valor = doc.to_dict().get("dados")
        return valor if valor is not None else default
    return default


def _salvar(chave: str, valor):
    get_db().collection("diagnostico").document(chave).set({
        "dados": valor,
        "atualizado_em": datetime.now(BRT).isoformat(),
    })


# ─────────────────────────────────────────────────────────────
# Lógica pura de consolidação automática (sem Streamlit/Firestore —
# fácil de testar isoladamente e reaproveitar).
# ─────────────────────────────────────────────────────────────
def normalizar_texto(t: str) -> str:
    t = (t or "").strip().lower()
    return re.sub(r"\s+", " ", t)


def calcular_duracao_min(hora_inicio: str, hora_fim: str) -> str:
    if not hora_inicio or not hora_fim:
        return ""
    try:
        h1, m1 = map(int, str(hora_inicio).split(":")[:2])
        h2, m2 = map(int, str(hora_fim).split(":")[:2])
    except Exception:
        return ""
    diff = (h2 * 60 + m2) - (h1 * 60 + m1)
    if diff < 0:
        diff += 24 * 60  # atravessou a meia-noite (raro, mas evita negativo)
    return str(diff) if diff > 0 else ""


def mais_frequente(valores) -> str:
    valores = [v for v in valores if v and str(v).strip()]
    if not valores:
        return ""
    contagem = {}
    for v in valores:
        contagem[v] = contagem.get(v, 0) + 1
    return max(contagem.items(), key=lambda kv: kv[1])[0]


def sugerir_matriz(diario, inventario, matriz_atual):
    """Gera linhas candidatas para a Matriz a partir do Diário de Bordo,
    pulando atividades que já existem na Matriz (por nome normalizado)."""
    ja_na_matriz = {normalizar_texto(r.get("atividade", "")) for r in matriz_atual}
    grupos = {}
    for linha in diario:
        ativ = (linha.get("atividade") or "").strip()
        if not ativ:
            continue
        chave = normalizar_texto(ativ)
        g = grupos.setdefault(chave, {"nome": ativ, "pessoas": [], "origens": [],
                                       "sistemas": [], "datas": [], "minutos": []})
        g["pessoas"].append(linha.get("analista", ""))
        g["origens"].append(linha.get("origem", ""))
        g["sistemas"].append(linha.get("sistemas", ""))
        g["datas"].append(linha.get("data", ""))
        dur = linha.get("duracao")
        if dur and str(dur).isdigit():
            g["minutos"].append(int(dur))

    registrada_por_atividade = {}
    for item in inventario:
        ativ = (item.get("atividade") or "").strip()
        if ativ:
            registrada_por_atividade[normalizar_texto(ativ)] = item.get("registrada", "")

    sugestoes = []
    for chave, g in grupos.items():
        if chave in ja_na_matriz:
            continue
        qtd_dias = len({d for d in g["datas"] if d})
        frequencia = "Diária" if qtd_dias >= 4 else ("Semanal" if qtd_dias >= 2 else "Eventual")
        tempo = ""
        if g["minutos"]:
            media = round(sum(g["minutos"]) / len(g["minutos"]))
            tempo = f"{media} min (média de {len(g['minutos'])} registro(s))"
        registrada = registrada_por_atividade.get(chave, "")
        visibilidade = ("Visível (mapeada)" if registrada == "Sim"
                        else ("Invisível (não mapeada)" if registrada else ""))
        sugestoes.append({
            "atividade": g["nome"], "quemfaz": mais_frequente(g["pessoas"]),
            "frequencia": frequencia, "tempo": tempo, "origem": mais_frequente(g["origens"]),
            "sistemas": mais_frequente(g["sistemas"]), "dependencia": "",
            "visibilidade": visibilidade, "percentual": "",
        })
    return sugestoes


def calcular_percentual_matriz(matriz):
    """% do tempo total da equipe = tempo médio ponderado pela frequência,
    normalizado a 100% entre as linhas com tempo preenchido. Aproximação
    de ordem de grandeza, não precisão contábil."""
    peso_frequencia = {"Diária": 5, "Semanal": 1, "Quinzenal": 0.5, "Mensal": 0.25, "Eventual": 0.3}
    esforcos = {}
    for m in matriz:
        ativ = (m.get("atividade") or "").strip()
        tempo = m.get("tempo") or ""
        if not ativ or not tempo:
            continue
        match = re.search(r"(\d+)", tempo)
        if not match:
            continue
        minutos = int(match.group(1))
        peso = peso_frequencia.get(m.get("frequencia", ""), 1)
        esforcos[normalizar_texto(ativ)] = minutos * peso
    total = sum(esforcos.values())
    return {k: (f"{round(v / total * 100, 1)}%" if total > 0 else "0%") for k, v in esforcos.items()}


def sugerir_respostas(inventario, organograma, diario, matriz, gemba):
    """Rascunho para cada uma das 6 perguntas finais, agregando as outras seções."""
    atividades = sorted({(i.get("atividade") or "").strip() for i in inventario if (i.get("atividade") or "").strip()})
    p1 = (f"A equipe realiza pelo menos {len(atividades)} tipo(s) de atividade registrados no inventário:\n- "
          + "\n- ".join(atividades)) if atividades else ""

    linhas_p2 = []
    for o in organograma:
        nome = (o.get("pessoa") or "").strip()
        if not nome:
            continue
        linha = f"{nome} — {o.get('papel', '')}"
        if o.get("especialista"):
            linha += f" (especialista em {o['especialista']})"
        if o.get("cobre"):
            linha += f"; cobre/é coberto por {o['cobre']}"
        linhas_p2.append(linha)
    p2 = "\n".join(linhas_p2)

    origens = [d.get("origem", "").strip() for d in diario if (d.get("origem") or "").strip()]
    p3 = ""
    if origens:
        total = len(origens)
        contagem = {}
        for o in origens:
            contagem[o] = contagem.get(o, 0) + 1
        p3 = "\n".join(f"{o}: {round(q / total * 100)}% dos registros do diário"
                       for o, q in sorted(contagem.items(), key=lambda kv: -kv[1]))

    tempos = {}
    for m in matriz:
        ativ = (m.get("atividade") or "").strip()
        tempo = m.get("tempo") or ""
        if not ativ or not tempo:
            continue
        match = re.search(r"(\d+)", tempo)
        if match:
            tempos[ativ] = int(match.group(1))
    top = sorted(tempos.items(), key=lambda kv: -kv[1])[:5]
    p4 = "\n".join(f"{a}: ~{t} min por ocorrência" for a, t in top)

    deps = [m.get("dependencia", "").strip() for m in matriz if (m.get("dependencia") or "").strip()]
    cobre = [o.get("cobre", "").strip() for o in organograma if (o.get("cobre") or "").strip()]
    p5 = "\n".join(dict.fromkeys(deps + cobre))

    def eh_invisivel(g):
        texto = ((g.get("insight") or "") + " " + (g.get("observado") or "")).lower()
        return "invisív" in texto or "não mapead" in texto or "nao mapead" in texto

    linhas_p6 = [(g.get("insight") or g.get("observado") or "").strip() for g in gemba if eh_invisivel(g)]
    linhas_p6 += [m.get("atividade", "").strip() for m in matriz if m.get("visibilidade") == "Invisível (não mapeada)"]
    p6 = "\n".join(dict.fromkeys([l for l in linhas_p6 if l]))

    return {"p1": p1, "p2": p2, "p3": p3, "p4": p4, "p5": p5, "p6": p6}


def sugerir_recursos_jornada(atividade, matriz, diario):
    """Sugere Responsável e Ferramentas para o Mapa de Jornada: olha a Matriz
    primeiro (mais confiável) e, na falta dela, agrega o Diário de Bordo bruto."""
    alvo = normalizar_texto(atividade)
    if not alvo:
        return {"responsavel": "", "ferramentas": ""}
    for m in matriz:
        if normalizar_texto(m.get("atividade", "")) == alvo:
            resp, ferr = (m.get("quemfaz") or "").strip(), (m.get("sistemas") or "").strip()
            if resp or ferr:
                return {"responsavel": resp, "ferramentas": ferr}
    relacionados = [d for d in diario if normalizar_texto(d.get("atividade", "")) == alvo]
    return {
        "responsavel": mais_frequente([d.get("analista", "") for d in relacionados]),
        "ferramentas": mais_frequente([d.get("sistemas", "") for d in relacionados]),
    }


def listar_todas_atividades(inventario, matriz, diario):
    nomes = set()
    for lista in (inventario, matriz):
        for item in lista:
            a = (item.get("atividade") or "").strip()
            if a:
                nomes.add(a)
    for d in diario:
        a = (d.get("atividade") or "").strip()
        if a:
            nomes.add(a)
    return sorted(nomes)


def sugerir_atividades_do_inventario(inventario, ja_existentes_normalizados):
    """Retorna os nomes de atividade do Inventário que ainda não aparecem em
    `ja_existentes_normalizados` (um set de nomes já normalizados via
    normalizar_texto). Usado para "carregar" atividades do Inventário em
    outras seções (RACI preliminar, Matriz RACI) sem duplicar nem sobrescrever
    o que já foi cadastrado manualmente ali."""
    nomes, vistos = [], set()
    for item in inventario:
        nome = (item.get("atividade") or "").strip()
        if not nome:
            continue
        norm = normalizar_texto(nome)
        if norm in ja_existentes_normalizados or norm in vistos:
            continue
        vistos.add(norm)
        nomes.append(nome)
    return nomes


# ─────────────────────────────────────────────────────────────
# Constantes (categorias, cores das etapas, indicadores)
# ─────────────────────────────────────────────────────────────
CATEGORIAS_INVENTARIO = ["Chamado oficial", "Apoio a outra área", "Retrabalho", "Reunião",
                          "Documentação", "Treinamento", "Outro"]
STATUS_REGISTRADA = ["Sim", "Não", "Parcialmente"]
CATEGORIAS_DIARIO = ["Atendimento reativo (chamado)", "Atividade proativa", "Reunião",
                     "E-mail / Comunicação", "Retrabalho", "Espera / Interrupção",
                     "Apoio a outra área (ex: N1)", "Controle paralelo (planilha)", "Outro"]
ORIGENS = ["N1", "Cliente direto", "Outra área", "E-mail", "Sistema", "Interno / próprio time"]
FREQUENCIAS = ["Diária", "Semanal", "Quinzenal", "Mensal", "Eventual"]
VISIBILIDADES = ["Visível (mapeada)", "Invisível (não mapeada)"]

# Matriz RACI "modelo planilha" (Fases/Atividades x Funções/Nomes) — mesma
# lista de opções e mesma lógica de grupos (Time Avaliado / Stakeholders)
# do arquivo de referência "Matriz de RACI - Diagnóstico.xlsx".
RACI_OPCOES = ["-", "R/C", "A/C", "I/A", "R/A", "I", "C", "A", "R"]
RACI_LEGENDA = [
    ("R", "Responsável", "#E53935"),
    ("A", "Aprovador", "#1E3A8A"),
    ("C", "Consultado", "#F2C230"),
    ("I", "Informado", "#2E7D32"),
]
GRUPOS_RACI = ["Time Avaliado", "Stakeholder"]

ETAPAS_JORNADA = [
    {"id": "e1", "nome": "1. RECEBER E ENTENDER A DEMANDA", "icone": "💬", "cor": "#1B2A4A"},
    {"id": "e2", "nome": "2. ANÁLISE E TRIAGEM",             "icone": "🔍", "cor": "#2F6FA8"},
    {"id": "e3", "nome": "3. TRATAR E RESOLVER",             "icone": "✅", "cor": "#2A9D8F"},
    {"id": "e4", "nome": "4. ACOMPANHAR E INFORMAR",         "icone": "📨", "cor": "#4CAF50"},
    {"id": "e5", "nome": "5. ENCERRAR E AVALIAR",            "icone": "📋", "cor": "#D9A83B"},
    {"id": "e6", "nome": "6. APRENDER E MELHORAR",           "icone": "📈", "cor": "#7B5EA7"},
]
LINHAS_JORNADA = [
    {"campo": "objetivo",    "label": "🎯 Objetivo da etapa",          "tipo": "textarea", "guia": "Objetivo desta etapa para esta atividade..."},
    {"campo": "acoes",       "label": "☰ Ações principais",            "tipo": "textarea", "guia": "- ação 1\n- ação 2"},
    {"campo": "ferramentas", "label": "💻 Ferramentas e sistemas",     "tipo": "text",     "guia": ""},
    {"campo": "responsavel", "label": "👤 Responsável",                "tipo": "text",     "guia": ""},
    {"campo": "experiencia", "label": "🙂 Experiência do colaborador", "tipo": "experiencia", "guia": ""},
    {"campo": "dores",       "label": "⚠️ Pontos de atenção (dores)", "tipo": "textarea", "guia": "- dor 1\n- dor 2"},
    {"campo": "melhorias",   "label": "💡 Oportunidades de melhoria", "tipo": "textarea", "guia": "- melhoria 1\n- melhoria 2"},
]
EMOJIS_EXPERIENCIA = ["😞", "😐", "🙂", "😀"]
INDICADORES = [
    ("TME", "Tempo Médio de Espera"), ("FRT", "Tempo Médio de 1ª Resposta"),
    ("TMA", "Tempo Médio de Atendimento"), ("Lead Time", "Tempo Total da Jornada"),
    ("SLA", "% dentro do Prazo"), ("FCR", "Resolução no 1º Contato"),
    ("Reabertura", "Taxa de Reabertura"), ("CSAT", "Satisfação do Cliente"),
    ("Reclame Aqui", "Índice de Solução/Reputação"),
]
ETAPAS_CHECKLIST = [
    ("inv", "Levantar o inventário de atividades", "14/07"),
    ("org", "Mapear o organograma funcional real da equipe (+ RACI)", "15–17/07"),
    ("dia", 'Aplicar o "Diário de Bordo"', "15–23/07"),
    ("ent", "Entrevista individual estruturada", "17–24/07"),
    ("gem", "Observação direta (Gemba) complementar", "22–24/07"),
    ("mat", "Consolidar a Matriz de Atividades x Responsabilidades x Tempo", "27–29/07"),
    ("jou", "Desenhar o Mapa de Jornada de Trabalho por atividade", "30–31/07"),
    ("val", "Validar com a equipe e a liderança", "03–04/08"),
]
PERGUNTAS = [
    ("p1", 'O que a equipe N2 faz (todos os tipos de atividade, não só os "chamados oficiais")?'),
    ("p2", "Quem faz o quê (responsabilidades individuais x coletivas, especialistas x generalistas)?"),
    ("p3", "Como o trabalho chega até eles (N1, direto do cliente, outras áreas, e-mail, sistema)?"),
    ("p4", "Quanto tempo cada atividade consome, na jornada real (não na teórica de 8h)?"),
    ("p5", "Onde estão as dependências externas (outras áreas, fornecedores, sistemas legados)?"),
    ("p6", "Onde está o trabalho invisível (não mapeado oficialmente, mas que toma tempo real)?"),
]


# ─────────────────────────────────────────────────────────────
# Helpers de conversão de tipo (Date/Time <-> string, p/ Firestore)
# ─────────────────────────────────────────────────────────────
def _to_date(v):
    if isinstance(v, date):
        return v
    if not v:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(v), fmt).date()
        except Exception:
            pass
    return None


def _date_to_str(v):
    return v.strftime("%Y-%m-%d") if isinstance(v, date) else (v or "")


def _to_time(v):
    if isinstance(v, dtime):
        return v
    if not v:
        return None
    try:
        h, m = map(int, str(v).split(":")[:2])
        return dtime(hour=h, minute=m)
    except Exception:
        return None


def _time_to_str(v):
    return v.strftime("%H:%M") if isinstance(v, dtime) else (v or "")


# ─────────────────────────────────────────────────────────────
# Editor genérico p/ tabelas simples (texto/select, sem campo calculado)
# ─────────────────────────────────────────────────────────────
def _editor_tabela_simples(chave, colunas_config, campos_ordem, pode_edit, editor_key):
    registros = _ler(chave, [])
    df = pd.DataFrame(registros) if registros else pd.DataFrame([{}])
    for c in campos_ordem:
        if c not in df.columns:
            df[c] = ""
    df = df[campos_ordem].fillna("")

    editado = st.data_editor(
        df, num_rows="dynamic", use_container_width=True, hide_index=True,
        key=editor_key, column_config=colunas_config, disabled=not pode_edit,
    ).fillna("")

    if pode_edit and not editado.equals(df):
        _salvar(chave, editado.to_dict("records"))
    return editado


# ─────────────────────────────────────────────────────────────
# 01 · Checklist
# ─────────────────────────────────────────────────────────────
def _tab_checklist():
    st.markdown("#### ✅ Checklist da metodologia")
    st.caption("Baseado no documento de diretrizes do mapeamento — janela de 14/07 a 05/08/2026.")
    marcados = _ler("checklist", {})
    for id_, titulo, datas in ETAPAS_CHECKLIST:
        valor = st.checkbox(f"**{titulo}** · `{datas}`", value=marcados.get(id_, False), key=f"diag_check_{id_}")
        if valor != marcados.get(id_, False):
            marcados[id_] = valor
            _salvar("checklist", marcados)

    with st.expander("📖 Premissas do levantamento"):
        st.markdown(
            "- O objetivo é compreender o **processo**, não avaliar desempenho individual.\n"
            "- Registre **tudo**: esperas, interrupções, atividades de apoio.\n"
            "- O levantamento deve refletir a rotina **real**, sem alterações de comportamento.\n"
            "- As informações são usadas só para melhoria contínua e otimização operacional."
        )


# ─────────────────────────────────────────────────────────────
# 02 · Inventário
# ─────────────────────────────────────────────────────────────
def _tab_inventario(pode_edit):
    st.markdown("#### 📋 Inventário de atividades")
    st.caption("Antes de qualquer entrevista: liste os tipos de atividade já conhecidos "
               "(tickets, tags do sistema, controles paralelos por fora do oficial).")
    _editor_tabela_simples(
        "inventario",
        {
            "atividade": st.column_config.TextColumn("Atividade", width="large"),
            "categoria": st.column_config.SelectboxColumn("Categoria", options=CATEGORIAS_INVENTARIO),
            "descricao": st.column_config.TextColumn("Descrição / exemplo", width="large"),
            "registrada": st.column_config.SelectboxColumn("Já registrada oficialmente?", options=STATUS_REGISTRADA),
        },
        ["atividade", "categoria", "descricao", "registrada"],
        pode_edit, "diag_editor_inventario",
    )


# ─────────────────────────────────────────────────────────────
# 03 · Organograma + RACI preliminar (por tipo de demanda)
# ─────────────────────────────────────────────────────────────
def _tab_organograma(pode_edit):
    st.markdown("#### 🧭 Organograma funcional real")
    st.caption("O organograma oficial nem sempre reflete quem faz o quê. Registre a divisão real: "
               "individuais x coletivas, especialistas x generalistas, hierarquia informal.")
    _editor_tabela_simples(
        "organograma",
        {
            "pessoa": st.column_config.TextColumn("Pessoa"),
            "papel": st.column_config.TextColumn("Papel real (não o cargo)"),
            "especialista": st.column_config.TextColumn("Especialista em"),
            "cobre": st.column_config.TextColumn("Cobre / é coberto por"),
            "obs": st.column_config.TextColumn("Observações", width="large"),
        },
        ["pessoa", "papel", "especialista", "cobre", "obs"],
        pode_edit, "diag_editor_organograma",
    )

    st.markdown("---")
    st.markdown("##### 🧩 RACI preliminar (por tipo de demanda)")
    st.caption("Visão rápida por tipo de demanda, recomendada na Etapa 2 do documento. "
               "Para a matriz completa (fases/atividades x pessoas, no modelo da planilha), use a aba **🧩 Matriz RACI**.")

    if pode_edit:
        raci_atual = _ler("raci", [])
        existentes = {normalizar_texto(r.get("tipo_demanda", "")) for r in raci_atual}
        novas = sugerir_atividades_do_inventario(_ler("inventario", []), existentes)
        if st.button("🔄 Carregar atividades do Inventário", key="diag_btn_raci_do_inventario"):
            if novas:
                linhas_novas = [{"tipo_demanda": n, "responsavel": "", "aprovador": "",
                                 "consultado": "", "informado": ""} for n in novas]
                _salvar("raci", raci_atual + linhas_novas)
                st.success(f"{len(linhas_novas)} atividade(s) carregada(s) do Inventário como tipo de demanda.")
                st.rerun()
            else:
                st.info("Nenhuma atividade nova encontrada no Inventário (ou já estão todas cadastradas aqui).")

    _editor_tabela_simples(
        "raci",
        {
            "tipo_demanda": st.column_config.TextColumn("Tipo de demanda", width="medium"),
            "responsavel": st.column_config.TextColumn("Responsável (R)"),
            "aprovador": st.column_config.TextColumn("Aprovador (A)"),
            "consultado": st.column_config.TextColumn("Consultado (C)"),
            "informado": st.column_config.TextColumn("Informado (I)"),
        },
        ["tipo_demanda", "responsavel", "aprovador", "consultado", "informado"],
        pode_edit, "diag_editor_raci",
    )


# ─────────────────────────────────────────────────────────────
# 04 · Matriz RACI (modelo da planilha) — Fases/Atividades x Funções/Nomes
# ─────────────────────────────────────────────────────────────
def resumo_raci_por_pessoa(matriz_raci, nomes):
    """Conta ocorrências de cada papel (R/A/C/I) por pessoa, considerando
    também combinações tipo 'R/C' (conta para os dois lados da combinação)."""
    linhas = []
    for nome in nomes:
        contagem = {"R": 0, "A": 0, "C": 0, "I": 0}
        for registro in matriz_raci:
            valor = str(registro.get(nome, "") or "").strip()
            if not valor or valor == "-":
                continue
            for letra in valor.split("/"):
                letra = letra.strip()
                if letra in contagem:
                    contagem[letra] += 1
        linhas.append({"Pessoa / Função": nome, **contagem})
    return linhas


def _exportar_raci_matriz_xlsx(projeto, data_str, pessoas, nomes, matriz_raci) -> bytes:
    """Gera o .xlsx no mesmo layout visual da planilha de referência: cabeçalho
    Data/Projeto, legenda R/A/C/I colorida, "Fases" à esquerda, colunas de
    pessoas agrupadas em "Time Avaliado" e "Stakeholders", e dropdown nas
    células de cruzamento com as mesmas opções do modelo original."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # Mesma ordem visual do modelo: todo o bloco "Time Avaliado" primeiro,
    # depois todo o bloco "Stakeholders" — por isso reordena em vez de manter
    # a ordem de cadastro (que pode intercalar os dois grupos).
    grupo_por_nome = {p.get("nome", "").strip(): p.get("grupo", "") for p in pessoas}
    nomes_time = [n for n in nomes if grupo_por_nome.get(n) != "Stakeholder"]
    nomes_stake = [n for n in nomes if grupo_por_nome.get(n) == "Stakeholder"]
    nomes_ordenados = nomes_time + nomes_stake

    linhas_dados = [m for m in matriz_raci if (m.get("atividade") or "").strip()]
    total_linhas = max(len(linhas_dados), 10)  # mantém ao menos 10 linhas, como no modelo
    total_cols = 2 + max(len(nomes_ordenados), 1)
    ultima_col_letra = get_column_letter(total_cols)
    primeira_linha_dados, ultima_linha_dados = 6, 6 + total_linhas - 1

    COR_VERMELHO, COR_AZUL, COR_AMARELO, COR_VERDE = "FFE53935", "FF1E3A8A", "FFF2C230", "FF2E7D32"
    COR_PRETO, COR_LAVANDA, COR_AZUL_CLARO, COR_CINZA, COR_CINZA_CLARO = (
        "FF000000", "FFD9D9FF", "FF6699FF", "FFBFBFBF", "FFE7E6E6")
    branco_bold = Font(color="FFFFFFFF", bold=True)
    preto_bold = Font(color="FF000000", bold=True)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda_fina = Border(*(Side(style="thin", color="FFBFBFBF"),) * 4)

    def preencher(coord, valor, fill=None, fonte=None, alinhar=True):
        celula = ws[coord]
        celula.value = valor
        if fill:
            celula.fill = PatternFill("solid", fgColor=fill)
        if fonte:
            celula.font = fonte
        if alinhar:
            celula.alignment = centro
        celula.border = borda_fina

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz RACI"

    # Linha 1 — Data / Nome do projeto
    ws.merge_cells("A1:B1")
    preencher("A1", f"Data: {data_str or '—'}", fill=COR_CINZA_CLARO, fonte=preto_bold)
    ws.merge_cells(f"C1:{ultima_col_letra}1")
    preencher("C1", f"Nome do Projeto:  {projeto or 'Diagnóstico N2'}", fonte=Font(bold=True, size=13), alinhar=False)
    ws["C1"].alignment = Alignment(horizontal="left", vertical="center")

    # Linhas 2-3 — legenda R/A/C/I + "Funções / Nome"
    preencher("A2", "R\n(Responsável)", fill=COR_VERMELHO, fonte=branco_bold)
    preencher("B2", "A\n(Aprovador)", fill=COR_AZUL, fonte=branco_bold)
    preencher("A3", "C\n(Consultado)", fill=COR_AMARELO, fonte=preto_bold)
    preencher("B3", "I\n(Informado)", fill=COR_VERDE, fonte=branco_bold)
    ws.merge_cells(f"C2:{ultima_col_letra}3")
    preencher("C2", "Funções / Nome", fill=COR_PRETO, fonte=Font(color="FFFFFFFF", bold=True, size=16))

    # Linha 4 — "Fases" + nome de cada pessoa/função (coluna a coluna)
    ws.merge_cells("A4:B4")
    preencher("A4", "Fases", fill=COR_LAVANDA, fonte=Font(bold=True, size=14))
    for idx, nome in enumerate(nomes_ordenados):
        col = get_column_letter(3 + idx)
        preencher(f"{col}4", nome, fill=COR_AZUL_CLARO, fonte=preto_bold)

    # Linha 5 — "Atividade" + "Time Avaliado" / "Stakeholders"
    ws.merge_cells("A5:B5")
    preencher("A5", "Atividade", fill=COR_PRETO, fonte=branco_bold)
    if nomes_time:
        c1, c2 = get_column_letter(3), get_column_letter(2 + len(nomes_time))
        if c1 != c2:
            ws.merge_cells(f"{c1}5:{c2}5")
        preencher(f"{c1}5", "Time Avaliado", fill=COR_AZUL, fonte=branco_bold)
    if nomes_stake:
        c1 = get_column_letter(3 + len(nomes_time))
        c2 = get_column_letter(2 + len(nomes_time) + len(nomes_stake))
        if c1 != c2:
            ws.merge_cells(f"{c1}5:{c2}5")
        preencher(f"{c1}5", "Stakeholders", fill=COR_AZUL, fonte=branco_bold)

    # Linhas 6+ — dados (ou linhas em branco prontas para preencher, como no modelo)
    for i in range(total_linhas):
        r = primeira_linha_dados + i
        ws.merge_cells(f"A{r}:B{r}")
        atividade = linhas_dados[i]["atividade"] if i < len(linhas_dados) else ""
        preencher(f"A{r}", atividade, fill=COR_LAVANDA if not atividade else None, alinhar=False)
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for idx, nome in enumerate(nomes_ordenados):
            col = get_column_letter(3 + idx)
            valor = (linhas_dados[i].get(nome, "-") or "-") if i < len(linhas_dados) else "-"
            preencher(f"{col}{r}", valor, fill=COR_CINZA)

    # Dropdown nas células de cruzamento — mesmas opções do modelo original
    dv = DataValidation(type="list", formula1='"-,R/C,A/C,I/A,R/A,I,C,A,R"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C{primeira_linha_dados}:{ultima_col_letra}{ultima_linha_dados}")

    # Larguras/alturas para ficar legível
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    for idx in range(len(nomes_ordenados)):
        ws.column_dimensions[get_column_letter(3 + idx)].width = 12
    for r in (2, 3):
        ws.row_dimensions[r].height = 30
    ws.row_dimensions[4].height = 22
    ws.freeze_panes = f"C{primeira_linha_dados}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _tab_raci_matriz(pode_edit):
    st.markdown("#### 🧩 Matriz RACI — Fases/Atividades x Funções")
    st.caption('Reproduz o modelo da planilha de referência: cabeçalho com projeto e data, '
               'colunas dinâmicas para "Time Avaliado" e "Stakeholders", e uma célula por '
               "cruzamento com as mesmas opções (-, R/C, A/C, I/A, R/A, I, C, A, R).")

    info = _ler("raci_matriz_info", {})
    c1, c2 = st.columns([2, 1])
    projeto = c1.text_input("Nome do projeto", value=info.get("projeto") or "Diagnóstico N2",
                             disabled=not pode_edit, key="diag_raci_projeto")
    data_ref = c2.date_input("Data", value=_to_date(info.get("data")) or date.today(),
                              disabled=not pode_edit, key="diag_raci_data")
    if pode_edit and (projeto != info.get("projeto") or _date_to_str(data_ref) != info.get("data")):
        _salvar("raci_matriz_info", {"projeto": projeto, "data": _date_to_str(data_ref)})

    st.markdown("&nbsp;", unsafe_allow_html=True)
    cols_legenda = st.columns(4)
    for c, (letra, desc, cor) in zip(cols_legenda, RACI_LEGENDA):
        c.markdown(
            f"<div style='background:{cor};color:#fff;border-radius:6px;padding:10px 6px;"
            f"text-align:center;font-weight:700;font-size:16px;'>{letra}"
            f"<div style='font-size:11px;font-weight:400;'>{desc}</div></div>",
            unsafe_allow_html=True,
        )
    st.caption("Combinações como **R/C**, **A/C**, **I/A**, **R/A** também são aceitas, "
               "como no modelo original, para casos em que o papel varia por sub-etapa.")

    st.markdown("---")
    st.markdown("##### 👥 Funções / Nomes (colunas da matriz)")
    st.caption('Cadastre aqui quem entra na matriz — separado em "Time Avaliado" (a própria '
               'célula N2) e "Stakeholders" (áreas/pessoas fora da célula).')
    pessoas = _ler("raci_pessoas", [])
    df_p = pd.DataFrame(pessoas) if pessoas else pd.DataFrame([{}])
    for c in ["nome", "grupo"]:
        if c not in df_p.columns:
            df_p[c] = ""
    df_p = df_p[["nome", "grupo"]].fillna("")

    editado_p = st.data_editor(
        df_p, num_rows="dynamic", use_container_width=True, hide_index=True,
        key="diag_editor_raci_pessoas",
        column_config={
            "nome": st.column_config.TextColumn("Nome / Função"),
            "grupo": st.column_config.SelectboxColumn("Grupo", options=GRUPOS_RACI),
        },
        disabled=not pode_edit,
    ).fillna("")

    if pode_edit and not editado_p.equals(df_p):
        _salvar("raci_pessoas", editado_p.to_dict("records"))

    # Usa o resultado já disponível em memória (editado_p) para montar as colunas da
    # matriz abaixo, em vez de reler do Firestore + forçar st.rerun(). Isso evita o
    # "piscar" de tela: um rerun forçado a cada edição, empilhado sobre o rerun que o
    # Streamlit já dispara sozinho sempre que um data_editor muda.
    pessoas = editado_p.to_dict("records")
    nomes = [p.get("nome", "").strip() for p in pessoas if (p.get("nome") or "").strip()]
    if not nomes:
        st.info("Cadastre ao menos uma pessoa/função acima para liberar a matriz de cruzamento.")
        return

    st.markdown("---")
    st.markdown("##### 📐 Matriz — Fases/Atividades x Funções")
    st.caption("Cada linha é uma fase ou atividade; cada coluna, uma pessoa/função cadastrada acima. "
               "Renomear ou remover uma pessoa não migra os valores já preenchidos para ela — "
               "ajuste a matriz em seguida se isso acontecer.")

    matriz_raci = _ler("raci_matriz", [])

    if pode_edit:
        existentes = {normalizar_texto(r.get("atividade", "")) for r in matriz_raci}
        novas = sugerir_atividades_do_inventario(_ler("inventario", []), existentes)
        if st.button("🔄 Carregar atividades do Inventário", key="diag_btn_raci_matriz_do_inventario"):
            if novas:
                linhas_novas = [{"atividade": n, **{nome: "-" for nome in nomes}} for n in novas]
                _salvar("raci_matriz", matriz_raci + linhas_novas)
                st.success(f"{len(linhas_novas)} atividade(s) carregada(s) do Inventário como linha da matriz.")
                st.rerun()
            else:
                st.info("Nenhuma atividade nova encontrada no Inventário (ou já estão todas na matriz).")
        matriz_raci = _ler("raci_matriz", [])  # relê (pode ter mudado pelo botão acima)

    colunas = ["atividade"] + nomes
    df_m = pd.DataFrame(matriz_raci) if matriz_raci else pd.DataFrame([{}])
    for c in colunas:
        if c not in df_m.columns:
            df_m[c] = "-"
    df_m = df_m[colunas].fillna("-")

    col_config = {"atividade": st.column_config.TextColumn("Fase / Atividade", width="large")}
    grupo_por_nome = {p.get("nome", "").strip(): p.get("grupo", "") for p in pessoas}
    for nome in nomes:
        grupo = grupo_por_nome.get(nome, "")
        icone = "🧑‍💼" if grupo == "Time Avaliado" else ("🤝" if grupo == "Stakeholder" else "")
        rotulo = f"{icone} {nome}".strip()
        col_config[nome] = st.column_config.SelectboxColumn(rotulo, options=RACI_OPCOES)

    editado_m = st.data_editor(
        df_m, num_rows="dynamic", use_container_width=True, hide_index=True,
        key="diag_editor_raci_matriz", column_config=col_config, disabled=not pode_edit,
    ).fillna("-")

    if pode_edit and not editado_m.equals(df_m):
        _salvar("raci_matriz", editado_m.to_dict("records"))

    with st.expander("📊 Resumo por pessoa/função"):
        resumo = resumo_raci_por_pessoa(editado_m.to_dict("records"), nomes)
        st.dataframe(pd.DataFrame(resumo), use_container_width=True, hide_index=True)
        st.caption("Contagem de quantas fases/atividades cada pessoa aparece como R, A, C ou I "
                   "(combinações como \"R/C\" contam para ambos os papéis).")

    c1, c2 = st.columns(2)
    with c1:
        try:
            xlsx_bytes = _exportar_raci_matriz_xlsx(projeto, _date_to_str(data_ref), pessoas, nomes,
                                                     editado_m.to_dict("records"))
            st.download_button(
                "⬇️ Baixar Matriz RACI (Excel, no modelo da planilha)", data=xlsx_bytes,
                file_name="matriz_raci_diagnostico_n2.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="diag_download_raci_xlsx", use_container_width=True,
            )
        except ImportError:
            st.caption("💡 Instale `openpyxl` (`pip install openpyxl`) para habilitar a exportação em Excel.")
    with c2:
        st.download_button(
            "⬇️ Baixar Matriz RACI (CSV)", data=editado_m.to_csv(index=False).encode("utf-8-sig"),
            file_name="matriz_raci_diagnostico_n2.csv", mime="text/csv", key="diag_download_raci_csv",
            use_container_width=True,
        )


# ─────────────────────────────────────────────────────────────
# 05 · Diário de Bordo
# ─────────────────────────────────────────────────────────────
def _tab_diario(pode_edit):
    st.markdown("#### 📓 Diário de Bordo")
    st.caption("Cada analista preenche uma linha por atividade, por 5 a 10 dias úteis. "
               "A duração é calculada automaticamente a partir da hora de início e fim.")

    registros = _ler("diario", [])
    colunas = ["data", "analista", "hora_inicio", "hora_fim", "duracao", "atividade",
               "categoria", "origem", "sistemas", "depende_terceiros", "quem", "obs"]
    df = pd.DataFrame(registros) if registros else pd.DataFrame([{}])
    for c in colunas:
        if c not in df.columns:
            df[c] = ""
    df = df[colunas].fillna("")
    df["data"] = df["data"].apply(_to_date)
    df["hora_inicio"] = df["hora_inicio"].apply(_to_time)
    df["hora_fim"] = df["hora_fim"].apply(_to_time)

    editado = st.data_editor(
        df, num_rows="dynamic", use_container_width=True, hide_index=True, key="diag_editor_diario",
        column_config={
            "data": st.column_config.DateColumn("Data", format="DD/MM/YYYY"),
            "analista": st.column_config.TextColumn("Analista"),
            "hora_inicio": st.column_config.TimeColumn("Hora início", format="HH:mm"),
            "hora_fim": st.column_config.TimeColumn("Hora fim", format="HH:mm"),
            "duracao": st.column_config.TextColumn("Duração (min)", disabled=True),
            "atividade": st.column_config.TextColumn("Atividade realizada", width="medium"),
            "categoria": st.column_config.SelectboxColumn("Categoria", options=CATEGORIAS_DIARIO),
            "origem": st.column_config.SelectboxColumn("Origem da demanda", options=ORIGENS),
            "sistemas": st.column_config.TextColumn("Sistema(s) utilizado(s)"),
            "depende_terceiros": st.column_config.SelectboxColumn("Depende de terceiros?", options=["Não", "Sim"]),
            "quem": st.column_config.TextColumn("Quem? (se sim)"),
            "obs": st.column_config.TextColumn("Observações", width="large"),
        },
        disabled=not pode_edit,
    )

    editado = editado.copy()
    editado["duracao"] = editado.apply(
        lambda r: calcular_duracao_min(_time_to_str(r["hora_inicio"]), _time_to_str(r["hora_fim"])), axis=1
    )

    if pode_edit:
        registros_novos = [{
            "data": _date_to_str(r["data"]), "analista": r["analista"] or "",
            "hora_inicio": _time_to_str(r["hora_inicio"]), "hora_fim": _time_to_str(r["hora_fim"]),
            "duracao": r["duracao"] or "", "atividade": r["atividade"] or "",
            "categoria": r["categoria"] or "", "origem": r["origem"] or "",
            "sistemas": r["sistemas"] or "", "depende_terceiros": r["depende_terceiros"] or "",
            "quem": r["quem"] or "", "obs": r["obs"] or "",
        } for _, r in editado.iterrows()]

        # Precisamos de UM rerun por edição real para a coluna "Duração" (calculada)
        # aparecer atualizada em tela. Sem a "assinatura" abaixo, uma pequena instabilidade
        # de tipo entre o que acabamos de montar e o que está salvo (ex.: data/hora indo e
        # voltando de string para objeto a cada execução) pode fazer essa comparação nunca
        # "estabilizar", disparando st.rerun() sem parar — daí a tela ficar piscando.
        # Guardando a assinatura do que já foi salvo NESTA sessão, o rerun só acontece
        # uma vez por alteração de verdade, mesmo que a comparação abaixo seja instável.
        assinatura = json.dumps(registros_novos, sort_keys=True, ensure_ascii=False)
        if registros_novos != registros and st.session_state.get("_diag_diario_assinatura") != assinatura:
            _salvar("diario", registros_novos)
            st.session_state["_diag_diario_assinatura"] = assinatura
            st.rerun()

    st.caption('💡 Registre **tudo**: reuniões, e-mails, retrabalho, espera por resposta de outra área, '
               'e atividades "não oficiais" (ajudar N1, apagar incêndio, planilha paralela). '
               "O objetivo é entender o processo, não avaliar desempenho individual.")


# ─────────────────────────────────────────────────────────────
# 06 · Entrevistas
# ─────────────────────────────────────────────────────────────
def _tab_entrevistas(pode_edit):
    st.markdown("#### 🎙️ Entrevista individual estruturada")
    with st.expander("📜 Roteiro sugerido (30–45 min por pessoa)"):
        perguntas = [
            "Descreva um dia de trabalho normal, do início ao fim.",
            "Quais atividades você faz que não aparecem no sistema de chamados?",
            "De onde vêm as demandas que você recebe (N1, cliente direto, outras áreas, e-mail, sistema)?",
            "Qual atividade consome mais tempo do que deveria? Por quê?",
            "De quais áreas, sistemas ou pessoas fora da equipe você depende pra concluir seu trabalho?",
            "O que você faz que ninguém mais da equipe sabe fazer?",
            "O que te interrompe com mais frequência durante o dia?",
            "Se pudesse eliminar uma atividade do seu dia, qual seria e por quê?",
        ]
        for i, p in enumerate(perguntas, 1):
            st.markdown(f"**{i}.** {p}")
        st.caption("Individual, não em grupo — em grupo há tendência de as pessoas se calarem "
                   "sobre dificuldades reais ou discrepâncias de carga de trabalho entre colegas.")

    registros = _ler("entrevistas", [])
    df = pd.DataFrame(registros) if registros else pd.DataFrame([{}])
    for c in ["nome", "data", "respostas"]:
        if c not in df.columns:
            df[c] = ""
    df = df[["nome", "data", "respostas"]].fillna("")
    df["data"] = df["data"].apply(_to_date)

    editado = st.data_editor(
        df, num_rows="dynamic", use_container_width=True, hide_index=True, key="diag_editor_entrevistas",
        column_config={
            "nome": st.column_config.TextColumn("Nome"),
            "data": st.column_config.DateColumn("Data da entrevista", format="DD/MM/YYYY"),
            "respostas": st.column_config.TextColumn("Anotações / respostas", width="large"),
        },
        disabled=not pode_edit,
    )
    if pode_edit:
        registros_novos = [{"nome": r["nome"] or "", "data": _date_to_str(r["data"]), "respostas": r["respostas"] or ""}
                           for _, r in editado.iterrows()]
        if registros_novos != registros:
            _salvar("entrevistas", registros_novos)


# ─────────────────────────────────────────────────────────────
# 07 · Gemba
# ─────────────────────────────────────────────────────────────
def _tab_gemba(pode_edit):
    st.markdown("#### 👣 Observação direta (Gemba)")
    st.info(
        '**Gemba (現場)** — termo japonês do Lean que significa "o lugar real", onde o trabalho de '
        "fato acontece. Observe ao vivo, sem interromper e sem corrigir nada durante a observação — "
        "o objetivo é ver o real, não o ideal."
    )
    _editor_tabela_simples(
        "gemba",
        {
            "data": st.column_config.TextColumn("Data"),
            "horario": st.column_config.TextColumn("Horário"),
            "pessoa": st.column_config.TextColumn("Pessoa observada"),
            "observado": st.column_config.TextColumn("O que foi observado", width="large"),
            "insight": st.column_config.TextColumn("Insight / trabalho invisível?", width="large"),
        },
        ["data", "horario", "pessoa", "observado", "insight"],
        pode_edit, "diag_editor_gemba",
    )


# ─────────────────────────────────────────────────────────────
# 08 · Matriz de Atividades x Responsabilidades x Tempo
# ─────────────────────────────────────────────────────────────
def _tab_matriz(pode_edit):
    st.markdown("#### 📐 Matriz de Atividades x Responsabilidades x Tempo")
    st.caption("Cruza tudo que foi coletado no Diário, nas entrevistas e no Gemba — "
               "colunas na ordem do documento de diretrizes (Etapa 6).")

    diario = _ler("diario", [])
    inventario = _ler("inventario", [])
    matriz = _ler("matriz", [])

    if pode_edit:
        c1, c2 = st.columns(2)
        with c1:
            if st.button("🔄 Gerar sugestões a partir do Diário", key="diag_btn_sugerir_matriz", use_container_width=True):
                novas = sugerir_matriz(diario, inventario, matriz)
                if novas:
                    _salvar("matriz", matriz + novas)
                    st.success(f"{len(novas)} linha(s) sugerida(s) adicionada(s). "
                               'Revise antes de finalizar (principalmente "dependência", que não é gerada automaticamente).')
                else:
                    st.info("Nenhuma atividade nova encontrada no Diário (ou já estão todas na Matriz).")
                st.rerun()
        with c2:
            if st.button("📐 Recalcular % do tempo total", key="diag_btn_recalcular_pct", use_container_width=True):
                pct = calcular_percentual_matriz(matriz)
                for linha in matriz:
                    norm = normalizar_texto(linha.get("atividade", ""))
                    if norm in pct:
                        linha["percentual"] = pct[norm]
                _salvar("matriz", matriz)
                st.rerun()

    colunas = ["atividade", "quemfaz", "frequencia", "tempo", "origem", "sistemas",
               "dependencia", "visibilidade", "percentual"]
    matriz = _ler("matriz", [])  # relê (pode ter mudado pelos botões acima)
    df = pd.DataFrame(matriz) if matriz else pd.DataFrame([{}])
    for c in colunas:
        if c not in df.columns:
            df[c] = ""
    df = df[colunas].fillna("")

    editado = st.data_editor(
        df, num_rows="dynamic", use_container_width=True, hide_index=True, key="diag_editor_matriz",
        column_config={
            "atividade": st.column_config.TextColumn("Atividade", width="medium"),
            "quemfaz": st.column_config.TextColumn("Quem executa"),
            "frequencia": st.column_config.SelectboxColumn("Frequência", options=FREQUENCIAS),
            "tempo": st.column_config.TextColumn("Tempo médio", help="ex: 25 min"),
            "origem": st.column_config.SelectboxColumn("Origem da demanda", options=ORIGENS),
            "sistemas": st.column_config.TextColumn("Sistemas usados"),
            "dependencia": st.column_config.TextColumn("Depende de terceiros?", help="ex: Sim - TI"),
            "visibilidade": st.column_config.SelectboxColumn("Visibilidade", options=VISIBILIDADES),
            "percentual": st.column_config.TextColumn("% do tempo total", disabled=True),
        },
        disabled=not pode_edit,
    ).fillna("")

    if pode_edit and not editado.equals(df):
        _salvar("matriz", editado.to_dict("records"))


# ─────────────────────────────────────────────────────────────
# 09 · Mapa de Jornada de Trabalho por Atividade
# ─────────────────────────────────────────────────────────────
def _etapa_vazia():
    campos = {c["campo"]: "" for c in LINHAS_JORNADA}
    campos["emoji"] = "🙂"
    return campos


def _jornada_vazia(atividade):
    return {"atividade": atividade, "etapas": {e["id"]: _etapa_vazia() for e in ETAPAS_JORNADA}}


def _indicadores_relacionados():
    st.markdown("**📊 Indicadores relacionados**")
    linha1, linha2 = INDICADORES[:5], INDICADORES[5:]
    for linha in (linha1, linha2):
        cols = st.columns(len(linha))
        for c, (sigla, desc) in zip(cols, linha):
            c.markdown(
                f"<div style='background:#1B2A4A;color:#fff;border-radius:6px;padding:8px 4px;"
                f"text-align:center;font-size:11px;margin-bottom:6px;'>"
                f"<b>{sigla}</b><br><span style='font-size:9px;opacity:.85;'>{desc}</span></div>",
                unsafe_allow_html=True,
            )


def _exportar_jornada_markdown(atividade, jornada) -> str:
    linhas = [f"# Mapa de Jornada de Trabalho — {atividade}\n"]
    for etapa in ETAPAS_JORNADA:
        et = jornada["etapas"][etapa["id"]]
        linhas.append(f"## {etapa['icone']} {etapa['nome']}\n")
        for linha in LINHAS_JORNADA:
            if linha["campo"] == "experiencia":
                linhas.append(f"**{linha['label']}:** {et.get('emoji', '')} {et.get('experiencia', '') or '—'}\n")
            else:
                linhas.append(f"**{linha['label']}:** {et.get(linha['campo'], '') or '—'}\n")
        linhas.append("")
    linhas.append("## 📊 Indicadores relacionados\n")
    for sigla, desc in INDICADORES:
        linhas.append(f"- **{sigla}** — {desc}")
    return "\n".join(linhas)


def _pdf_safe(texto: str) -> str:
    """Remove emojis e outros caracteres fora do Latin-1 — as fontes core
    do fpdf2 (Helvetica) não suportam Unicode; acentos em português ficam
    intactos porque Latin-1 os cobre. Travessões/aspas tipográficas viram
    equivalentes ASCII antes de remover o resto (senão somem sem deixar rastro)."""
    texto = texto or ""
    for original, troca in {"—": "-", "–": "-", "’": "'", "‘": "'", "“": '"', "”": '"', "…": "..."}.items():
        texto = texto.replace(original, troca)
    return texto.encode("latin-1", "ignore").decode("latin-1")


def _exportar_jornada_pdf(atividade, jornada) -> bytes:
    """Exportação em PDF, best-effort: um bloco por etapa (empilhado), não
    lado a lado como no modelo visual de referência — trade-off para manter
    a geração simples e robusta com fpdf2. Levanta ImportError se o fpdf2
    não estiver instalado (o chamador deve tratar isso)."""
    from fpdf import FPDF, XPos, YPos  # import local: só é exigido se o botão for usado

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.multi_cell(0, 10, _pdf_safe(f"Mapa de Jornada de Trabalho — {atividade}"),
                   new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    cores_rgb = {
        "e1": (27, 42, 74), "e2": (47, 111, 168), "e3": (42, 157, 143),
        "e4": (76, 175, 80), "e5": (217, 168, 59), "e6": (123, 94, 167),
    }
    for etapa in ETAPAS_JORNADA:
        et = jornada["etapas"][etapa["id"]]
        r, g, b = cores_rgb[etapa["id"]]
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 9, _pdf_safe(etapa["nome"]), new_x=XPos.LMARGIN, new_y=YPos.NEXT, fill=True)
        pdf.set_text_color(20, 20, 20)
        for linha in LINHAS_JORNADA:
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(0, 6, _pdf_safe(linha["label"]), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font("Helvetica", "", 9)
            if linha["campo"] == "experiencia":
                texto = f"{et.get('experiencia', '') or '(sem preenchimento)'}"
            else:
                texto = et.get(linha["campo"], "") or "(sem preenchimento)"
            pdf.multi_cell(0, 6, _pdf_safe(texto), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(1)
        pdf.ln(3)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 8, "Indicadores relacionados", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 9)
    for sigla, desc in INDICADORES:
        pdf.multi_cell(0, 5, _pdf_safe(f"{sigla} - {desc}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    saida = pdf.output()
    return bytes(saida)


def _tab_jornada(pode_edit):
    st.markdown("#### 🗺️ Mapa de Jornada de Trabalho — por atividade")
    st.caption("Para cada atividade/tipo de demanda: objetivo, ações, ferramentas, responsável, "
               "experiência do colaborador, dores e oportunidades de melhoria por etapa — "
               "no formato do modelo de referência da área.")

    inventario = _ler("inventario", [])
    matriz = _ler("matriz", [])
    diario = _ler("diario", [])
    jornadas = _ler("jornadas", {})

    atividades = listar_todas_atividades(inventario, matriz, diario)
    opcoes = ["— selecione —"] + atividades + ["+ nova atividade..."]
    escolha = st.selectbox("Atividade", opcoes, key="diag_jornada_select")

    atividade = None
    if escolha == "+ nova atividade...":
        nova = st.text_input("Nome da nova atividade", key="diag_jornada_nova", disabled=not pode_edit)
        if nova.strip():
            atividade = nova.strip()
    elif escolha != "— selecione —":
        atividade = escolha

    if not atividade:
        st.info("Escolha ou digite uma atividade para ver/editar o mapa de jornada dela. "
                 "Cada atividade tem seu próprio mapa salvo.")
        return

    norm = normalizar_texto(atividade)
    jornada = jornadas.get(norm) or _jornada_vazia(atividade)

    if pode_edit and st.button("🔄 Sugerir Responsável/Ferramentas (Matriz/Diário)",
                                key="diag_btn_sugerir_recursos"):
        rec = sugerir_recursos_jornada(atividade, matriz, diario)
        preenchidos = 0
        for e in ETAPAS_JORNADA:
            et = jornada["etapas"][e["id"]]
            if not et.get("responsavel") and rec["responsavel"]:
                et["responsavel"] = rec["responsavel"]
                st.session_state[f"diag_j_{norm}_{e['id']}_responsavel"] = rec["responsavel"]
                preenchidos += 1
            if not et.get("ferramentas") and rec["ferramentas"]:
                et["ferramentas"] = rec["ferramentas"]
                st.session_state[f"diag_j_{norm}_{e['id']}_ferramentas"] = rec["ferramentas"]
                preenchidos += 1
        jornadas[norm] = jornada
        _salvar("jornadas", jornadas)
        if preenchidos:
            st.success(f"{preenchidos} campo(s) preenchido(s). Ajuste por etapa se fizer sentido "
                       "(ex: quem recebe pode não ser quem resolve).")
        else:
            st.info("Ainda não há dados suficientes na Matriz/Diário para esta atividade.")
        st.rerun()

    st.markdown(f"##### Mapa de Jornada — **{atividade}**")

    cols_header = st.columns(6)
    for c, etapa in zip(cols_header, ETAPAS_JORNADA):
        c.markdown(
            f"<div style='background:{etapa['cor']};color:#fff;padding:8px 6px;border-radius:6px;"
            f"text-align:center;font-size:11.5px;font-weight:700;min-height:58px;"
            f"display:flex;align-items:center;justify-content:center;'>"
            f"{etapa['icone']} {etapa['nome']}</div>", unsafe_allow_html=True,
        )

    alterado = False
    for linha in LINHAS_JORNADA:
        st.markdown(f"**{linha['label']}**")
        cols = st.columns(6)
        for c, etapa in zip(cols, ETAPAS_JORNADA):
            et = jornada["etapas"][etapa["id"]]
            key_base = f"diag_j_{norm}_{etapa['id']}_{linha['campo']}"
            if linha["tipo"] == "textarea":
                novo = c.text_area("etapa", value=et.get(linha["campo"], ""), key=key_base,
                                    placeholder=linha["guia"], height=100, disabled=not pode_edit,
                                    label_visibility="collapsed")
            elif linha["tipo"] == "experiencia":
                emoji_atual = et.get("emoji", "🙂")
                idx = EMOJIS_EXPERIENCIA.index(emoji_atual) if emoji_atual in EMOJIS_EXPERIENCIA else 2
                emoji = c.selectbox("emoji", EMOJIS_EXPERIENCIA, index=idx, key=key_base + "_emoji",
                                     disabled=not pode_edit, label_visibility="collapsed")
                novo = c.text_input("frase", value=et.get("experiencia", ""), key=key_base,
                                     placeholder="frase curta...", disabled=not pode_edit,
                                     label_visibility="collapsed")
                if emoji != et.get("emoji"):
                    et["emoji"] = emoji
                    alterado = True
            else:
                novo = c.text_input("etapa", value=et.get(linha["campo"], ""), key=key_base,
                                     disabled=not pode_edit, label_visibility="collapsed")
            if novo != et.get(linha["campo"], ""):
                et[linha["campo"]] = novo
                alterado = True

    if alterado and pode_edit:
        jornadas[norm] = jornada
        _salvar("jornadas", jornadas)

    st.markdown("---")
    _indicadores_relacionados()

    st.markdown("---")
    md = _exportar_jornada_markdown(atividade, jornada)
    st.download_button("⬇️ Baixar Mapa de Jornada (Markdown)", data=md,
                        file_name=f"mapa_jornada_{norm.replace(' ', '_')}.md", mime="text/markdown",
                        key="diag_download_jornada_md")
    try:
        pdf_bytes = _exportar_jornada_pdf(atividade, jornada)
        st.download_button("⬇️ Baixar Mapa de Jornada (PDF)", data=pdf_bytes,
                            file_name=f"mapa_jornada_{norm.replace(' ', '_')}.pdf", mime="application/pdf",
                            key="diag_download_jornada_pdf")
    except ImportError:
        st.caption("💡 Instale `fpdf2` (`pip install fpdf2`) para habilitar exportação em PDF "
                   "além do Markdown (o PDF sai em blocos empilhados, não lado a lado como o modelo visual).")


# ─────────────────────────────────────────────────────────────
# 10 · As 6 perguntas finais
# ─────────────────────────────────────────────────────────────
def _tab_respostas(pode_edit):
    st.markdown("#### ❓ As 6 perguntas que a entrega precisa responder")
    st.caption("Preencha à medida que os dados forem chegando. É o resumo executivo pra apresentar "
               "na validação (03–04/08) e entregar em 05/08.")

    if pode_edit and st.button("🔄 Preencher rascunho automático", key="diag_btn_sugerir_respostas"):
        respostas = _ler("respostas", {})
        sugestoes = sugerir_respostas(_ler("inventario", []), _ler("organograma", []),
                                       _ler("diario", []), _ler("matriz", []), _ler("gemba", []))
        preenchidas = 0
        for pid, _ in PERGUNTAS:
            atual = (respostas.get(pid) or "").strip()
            sugerida = (sugestoes.get(pid) or "").strip()
            if not atual and sugerida:
                respostas[pid] = sugerida
                st.session_state[f"diag_resposta_{pid}"] = sugerida
                preenchidas += 1
        _salvar("respostas", respostas)
        if preenchidas:
            st.success(f"{preenchidas} pergunta(s) preenchida(s). As que já tinham texto seu não foram tocadas.")
        else:
            st.info("Nenhuma pergunta em branco, ou ainda não há dados suficientes coletados.")
        st.rerun()

    respostas = _ler("respostas", {})
    for pid, texto in PERGUNTAS:
        valor = st.text_area(texto, value=respostas.get(pid, ""), key=f"diag_resposta_{pid}",
                              height=110, disabled=not pode_edit)
        if pode_edit and valor != respostas.get(pid, ""):
            respostas[pid] = valor
            _salvar("respostas", respostas)


# ─────────────────────────────────────────────────────────────
# 11 · Resumo com IA (opcional)
# ─────────────────────────────────────────────────────────────
def _gerar_resumo_ia():
    import requests

    chave = st.secrets.get("anthropic_api_key", "")
    if not chave:
        return None, "Configure `anthropic_api_key` em `st.secrets` para usar esta função (chave própria da Anthropic — não é a mesma conta do chat)."

    dados = {
        "inventario": _ler("inventario", []), "organograma": _ler("organograma", []),
        "raci": _ler("raci", []), "raci_pessoas": _ler("raci_pessoas", []),
        "raci_matriz": _ler("raci_matriz", []), "diario": _ler("diario", []),
        "entrevistas": _ler("entrevistas", []), "gemba": _ler("gemba", []),
        "matriz": _ler("matriz", []), "jornadas": _ler("jornadas", {}),
        "respostas": _ler("respostas", {}),
    }
    prompt = (
        "Você está ajudando a consolidar um mapeamento de atividades (estilo Lean/Gemba) de uma "
        "equipe de backoffice N2. Abaixo estão os dados brutos coletados, em JSON. Escreva um RESUMO "
        "EXECUTIVO em português, objetivo e assertivo, cobrindo nesta ordem: 1) o que a equipe faz de "
        "fato; 2) como o trabalho é dividido; 3) de onde vêm as demandas; 4) onde o tempo é mais "
        "consumido; 5) principais dependências externas; 6) trabalho invisível identificado; e termine "
        "com 3 a 5 recomendações práticas priorizadas. Não invente dados que não estejam no JSON — "
        "se uma seção estiver vazia, apenas mencione que precisa de mais coleta ali.\n\n"
        f"DADOS COLETADOS (JSON):\n{json.dumps(dados, ensure_ascii=False)}"
    )
    try:
        resposta = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json", "x-api-key": chave, "anthropic-version": "2023-06-01"},
            json={"model": "claude-sonnet-4-6", "max_tokens": 2000,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=60,
        )
    except requests.RequestException as e:
        return None, f"Erro de conexão com a API: {e}"

    if resposta.status_code != 200:
        try:
            msg = resposta.json().get("error", {}).get("message", f"HTTP {resposta.status_code}")
        except Exception:
            msg = f"HTTP {resposta.status_code}"
        return None, f"Erro da API Anthropic: {msg}"

    corpo = resposta.json()
    texto = "".join(b.get("text", "") for b in corpo.get("content", []) if b.get("type") == "text")
    return texto, None


def _tab_resumo_ia(pode_edit):
    st.markdown("#### 🤖 Resumo com IA (opcional)")
    st.caption("Envia os dados coletados para a API da Anthropic (Claude) gerar um resumo executivo "
               "já lapidado, editável depois. Precisa de `anthropic_api_key` em `st.secrets` — chave "
               "separada da usada no chat, com custo por chamada. O sistema funciona 100% sem isso.")

    if pode_edit and st.button("🤖 Gerar resumo executivo com IA", key="diag_btn_resumo_ia"):
        with st.spinner("Consultando a IA..."):
            texto, erro = _gerar_resumo_ia()
        if erro:
            st.error(erro)
        else:
            _salvar("resumo_ia", texto)
            st.session_state["diag_resumo_ia_area"] = texto
            st.success("Resumo gerado.")

    valor_salvo = _ler("resumo_ia", "")
    resumo = st.text_area("Resumo (editável)", value=st.session_state.get("diag_resumo_ia_area", valor_salvo),
                           height=320, key="diag_resumo_ia_area", disabled=not pode_edit)
    if pode_edit and resumo != valor_salvo:
        _salvar("resumo_ia", resumo)


# ─────────────────────────────────────────────────────────────
# 12 · Relatório Consolidado — extrai e organiza tudo que foi coletado
# ─────────────────────────────────────────────────────────────
def gerar_relatorio_consolidado(checklist, inventario, organograma, raci_simples, raci_info,
                                 raci_pessoas, raci_matriz, diario, entrevistas, gemba,
                                 matriz, jornadas, respostas, resumo_ia) -> str:
    """Percorre TODAS as seções na mesma ordem lógica da metodologia (Etapas 1 a 8) e
    monta um único documento — pensado para a validação (Etapa 8) e para a entrega final,
    sem depender da IA (a IA, quando usada, entra como seção extra ao final)."""
    linhas = ["# Diagnóstico N2 — Relatório Consolidado\n"]
    linhas.append(f"**Projeto:** {raci_info.get('projeto') or 'Diagnóstico N2'}  ")
    linhas.append(f"**Data de referência:** {raci_info.get('data') or '—'}\n")

    linhas.append("## 1. Andamento da metodologia\n")
    concluidas = sum(1 for id_, *_ in ETAPAS_CHECKLIST if checklist.get(id_))
    linhas.append(f"_{concluidas} de {len(ETAPAS_CHECKLIST)} etapas concluídas._\n")
    for id_, titulo, datas in ETAPAS_CHECKLIST:
        marcado = "✅" if checklist.get(id_) else "⬜"
        linhas.append(f"- {marcado} {titulo} (`{datas}`)")
    linhas.append("")

    linhas.append("## 2. Inventário de atividades\n")
    itens = [i for i in inventario if (i.get("atividade") or "").strip()]
    if itens:
        for i in itens:
            linhas.append(f"- **{i['atividade']}** ({i.get('categoria') or '—'}) "
                           f"— registrada oficialmente: {i.get('registrada') or '—'}")
    else:
        linhas.append("_Nenhuma atividade cadastrada ainda._")
    linhas.append("")

    linhas.append("## 3. Organograma funcional real\n")
    pessoas_org = [o for o in organograma if (o.get("pessoa") or "").strip()]
    if pessoas_org:
        for o in pessoas_org:
            linha = f"- **{o['pessoa']}** — {o.get('papel') or '—'}"
            if o.get("especialista"):
                linha += f"; especialista em {o['especialista']}"
            if o.get("cobre"):
                linha += f"; cobre/é coberto por {o['cobre']}"
            linhas.append(linha)
    else:
        linhas.append("_Organograma real ainda não levantado._")
    linhas.append("")

    linhas.append("## 4. RACI preliminar por tipo de demanda\n")
    raci_itens = [r for r in raci_simples if (r.get("tipo_demanda") or "").strip()]
    if raci_itens:
        for r in raci_itens:
            linhas.append(f"- **{r['tipo_demanda']}** — R: {r.get('responsavel') or '—'} | "
                           f"A: {r.get('aprovador') or '—'} | C: {r.get('consultado') or '—'} | "
                           f"I: {r.get('informado') or '—'}")
    else:
        linhas.append("_RACI preliminar ainda não preenchido._")
    linhas.append("")

    linhas.append("## 5. Matriz RACI (Fases/Atividades x Funções)\n")
    nomes = [p.get("nome", "").strip() for p in raci_pessoas if (p.get("nome") or "").strip()]
    linhas_matriz_raci = [m for m in raci_matriz if (m.get("atividade") or "").strip()]
    if nomes and linhas_matriz_raci:
        linhas.append("| Atividade | " + " | ".join(nomes) + " |")
        linhas.append("|" + "---|" * (len(nomes) + 1))
        for r in linhas_matriz_raci:
            linhas.append(f"| {r['atividade']} | " + " | ".join(str(r.get(n) or "-") for n in nomes) + " |")
    else:
        linhas.append("_Matriz RACI ainda não preenchida._")
    linhas.append("")

    linhas.append("## 6. Diário de Bordo — visão agregada\n")
    diario_valido = [d for d in diario if (d.get("atividade") or "").strip()]
    if diario_valido:
        linhas.append(f"- Total de registros: **{len(diario_valido)}**")
        minutos_por_categoria = {}
        for d in diario_valido:
            cat = d.get("categoria") or "Sem categoria"
            dur = d.get("duracao")
            if dur and str(dur).isdigit():
                minutos_por_categoria[cat] = minutos_por_categoria.get(cat, 0) + int(dur)
        for cat, mins in sorted(minutos_por_categoria.items(), key=lambda kv: -kv[1]):
            linhas.append(f"- {cat}: {mins} min (~{round(mins / 60, 1)}h)")
    else:
        linhas.append("_Diário de bordo ainda não preenchido._")
    linhas.append("")

    linhas.append("## 7. Entrevistas realizadas\n")
    feitas = [e for e in entrevistas if (e.get("nome") or "").strip()]
    if feitas:
        for e in feitas:
            linhas.append(f"- {e['nome']} — {e.get('data') or '—'}")
    else:
        linhas.append("_Nenhuma entrevista registrada ainda._")
    linhas.append("")

    linhas.append("## 8. Observações de campo (Gemba)\n")
    gemba_validos = [g for g in gemba if (g.get("observado") or "").strip()]
    if gemba_validos:
        for g in gemba_validos:
            linha = f"- {g.get('data') or ''} {g.get('horario') or ''} — {g.get('pessoa') or '—'}: {g['observado']}"
            if g.get("insight"):
                linha += f" _(insight: {g['insight']})_"
            linhas.append(linha)
    else:
        linhas.append("_Nenhuma observação de campo registrada ainda._")
    linhas.append("")

    linhas.append("## 9. Matriz de Atividades x Responsabilidades x Tempo\n")
    matriz_valida = [m for m in matriz if (m.get("atividade") or "").strip()]
    if matriz_valida:
        for m in matriz_valida:
            linhas.append(f"- **{m['atividade']}** — {m.get('quemfaz') or '—'} | "
                           f"{m.get('frequencia') or '—'} | {m.get('tempo') or '—'} | "
                           f"{m.get('percentual') or '—'} do tempo total")
    else:
        linhas.append("_Matriz de Atividades ainda não consolidada._")
    linhas.append("")

    linhas.append("## 10. Mapas de Jornada por atividade\n")
    if jornadas:
        for norm, j in jornadas.items():
            linhas.append(f"- **{j.get('atividade', norm)}** — mapa de jornada preenchido")
    else:
        linhas.append("_Nenhum mapa de jornada criado ainda._")
    linhas.append("")

    linhas.append("## 11. Respostas às 6 perguntas-chave\n")
    for pid, texto in PERGUNTAS:
        resposta = (respostas.get(pid) or "").strip()
        linhas.append(f"**{texto}**\n\n{resposta or '_(em aberto)_'}\n")

    if (resumo_ia or "").strip():
        linhas.append("## 12. Resumo executivo (gerado com IA)\n")
        linhas.append(resumo_ia.strip())

    return "\n".join(linhas)


def _exportar_relatorio_pdf(md_texto: str) -> bytes:
    """Conversão best-effort de Markdown simples para PDF (títulos, tabelas
    e listas), no mesmo espírito de _exportar_jornada_pdf: robusto e legível,
    não uma renderização Markdown completa."""
    from fpdf import FPDF, XPos, YPos

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    for linha in md_texto.split("\n"):
        texto = _pdf_safe(linha)
        if texto.startswith("# "):
            pdf.set_font("Helvetica", "B", 16)
            pdf.multi_cell(0, 10, texto[2:], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        elif texto.startswith("## "):
            pdf.ln(2)
            pdf.set_font("Helvetica", "B", 13)
            pdf.multi_cell(0, 8, texto[3:], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        elif texto.strip().startswith("|"):
            pdf.set_font("Courier", "", 7)
            pdf.multi_cell(0, 4, texto, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        elif texto.strip().startswith("**") and texto.strip().endswith("**"):
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(0, 5, texto.replace("**", ""), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        elif texto.strip().startswith("-"):
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 5, texto, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        elif texto.strip() == "":
            pdf.ln(2)
        else:
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 5, texto.replace("**", ""), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return bytes(pdf.output())


def _tab_relatorio_final(pode_edit):
    st.markdown("#### 📄 Relatório Consolidado")
    st.caption("Reúne, na mesma ordem lógica da metodologia, tudo o que foi coletado em todas as "
               "abas — pronto para a validação com a equipe/liderança (Etapa 8) e para a entrega final.")

    md = gerar_relatorio_consolidado(
        _ler("checklist", {}), _ler("inventario", []), _ler("organograma", []),
        _ler("raci", []), _ler("raci_matriz_info", {}), _ler("raci_pessoas", []),
        _ler("raci_matriz", []), _ler("diario", []), _ler("entrevistas", []),
        _ler("gemba", []), _ler("matriz", []), _ler("jornadas", {}),
        _ler("respostas", {}), _ler("resumo_ia", ""),
    )

    with st.expander("👁️ Pré-visualização", expanded=True):
        st.markdown(md)

    c1, c2 = st.columns(2)
    with c1:
        st.download_button("⬇️ Baixar Relatório (Markdown)", data=md,
                            file_name="relatorio_consolidado_diagnostico_n2.md", mime="text/markdown",
                            key="diag_download_relatorio_md", use_container_width=True)
    with c2:
        try:
            pdf_bytes = _exportar_relatorio_pdf(md)
            st.download_button("⬇️ Baixar Relatório (PDF)", data=pdf_bytes,
                                file_name="relatorio_consolidado_diagnostico_n2.pdf", mime="application/pdf",
                                key="diag_download_relatorio_pdf", use_container_width=True)
        except ImportError:
            st.caption("💡 Instale `fpdf2` (`pip install fpdf2`) para habilitar a exportação em PDF.")


# ─────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────
def renderizar_diagnostico(papel, user=None):
    editar = pode_editar(user) if user is not None else papel in ("adm", "supervisor")

    st.subheader("🗺️ Diagnóstico N2 — Mapeamento de Atividades Operacionais")
    st.caption("CX · Backoffice (N2) · Área responsável: PQI · entrega prevista 05/08/2026")
    if not editar:
        st.info("Modo somente leitura — peça a um supervisor ou administrador para editar.")

    abas = st.tabs([
        "✅ Checklist", "📋 Inventário", "🧭 Organograma", "🧩 Matriz RACI", "📓 Diário de Bordo",
        "🎙️ Entrevistas", "👣 Gemba", "📐 Matriz", "🗺️ Mapa de Jornada",
        "❓ 6 Perguntas", "🤖 Resumo IA", "📄 Relatório Final",
    ])
    with abas[0]:
        _tab_checklist()
    with abas[1]:
        _tab_inventario(editar)
    with abas[2]:
        _tab_organograma(editar)
    with abas[3]:
        _tab_raci_matriz(editar)
    with abas[4]:
        _tab_diario(editar)
    with abas[5]:
        _tab_entrevistas(editar)
    with abas[6]:
        _tab_gemba(editar)
    with abas[7]:
        _tab_matriz(editar)
    with abas[8]:
        _tab_jornada(editar)
    with abas[9]:
        _tab_respostas(editar)
    with abas[10]:
        _tab_resumo_ia(editar)
    with abas[11]:
        _tab_relatorio_final(editar)
